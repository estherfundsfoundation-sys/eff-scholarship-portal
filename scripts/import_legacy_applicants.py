"""Idempotently stage the approved Name Your Need 2026 workbook without logging PII."""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

WORKBOOK = Path(r"C:\Users\Shayna\Downloads\EFF_Name_Your_Need_Scholarshi2026-07-14_16_05_07.xlsx")
PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = PROJECT_ROOT / ".env.production.local" if (PROJECT_ROOT / ".env.production.local").exists() else PROJECT_ROOT / ".env.local"
EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def load_env() -> dict[str, str]:
    values: dict[str, str] = {}
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        if line and not line.lstrip().startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            values[key.strip()] = value.strip().strip('"').strip("'")
    return values


class ApiResponse:
    def __init__(self, body: bytes): self.body = body
    def json(self): return json.loads(self.body.decode("utf-8")) if self.body else None


def api(method: str, path: str, env: dict[str, str], payload: Any | None = None, params: dict[str, str] | None = None, prefer: str | None = None) -> ApiResponse:
    headers = {"apikey": env["SUPABASE_SERVICE_ROLE_KEY"], "Authorization": f"Bearer {env['SUPABASE_SERVICE_ROLE_KEY']}", "Content-Type": "application/json"}
    if prefer:
        headers["Prefer"] = prefer
    query = f"?{urllib.parse.urlencode(params)}" if params else ""
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    request = urllib.request.Request(f"{env['NEXT_PUBLIC_SUPABASE_URL']}/rest/v1/{path}{query}", data=data, headers=headers, method=method)
    with urllib.request.urlopen(request, timeout=60) as response:
        return ApiResponse(response.read())


def chunks(items: list[dict[str, Any]], size: int = 200):
    for start in range(0, len(items), size):
        yield items[start:start + size]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--commit", action="store_true")
    parser.add_argument("--export", type=Path)
    args = parser.parse_args()
    digest = hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
    sheet = load_workbook(WORKBOOK, read_only=True, data_only=True).active
    source_headers = [scalar(cell.value) or "No Label" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    unique_headers = [f"{header} [{index}]" if source_headers.count(header) > 1 else header for index, header in enumerate(source_headers, 1)]
    parsed: list[dict[str, Any]] = []
    seen: Counter[str] = Counter()
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        cells = [scalar(value) for value in row]
        email = cells[8].lower()
        valid = bool(EMAIL_RE.match(email))
        duplicate = seen[email] > 0 if valid else False
        if valid:
            seen[email] += 1
        categories = []
        for index in (24, 25, 26, 27):
            categories.extend(part.strip() for part in cells[index].splitlines() if part.strip())
        normalized = {
            "legal_name": cells[1], "preferred_name": cells[2], "date_of_birth": cells[3], "gender": cells[4],
            "race_ethnicity": cells[5], "marital_status": cells[6], "residency_status": cells[7], "personal_email": email,
            "school_email": cells[9], "phone": cells[10], "address": cells[11], "institution": cells[12], "student_id": cells[13],
            "class_standing": cells[14], "major": cells[15], "expected_graduation": cells[16], "enrollment_status": cells[17], "gpa": cells[18],
            "emergency_contact_name": cells[19], "emergency_contact_relationship": cells[20], "emergency_contact_phone": cells[21], "emergency_contact_email": cells[22],
            "amount_requested": cells[23], "need_category": " | ".join(dict.fromkeys(categories)), "financial_need_description": cells[28],
            "story": cells[29], "faith_reflection": cells[30], "headshot_url": cells[31], "enrollment_proof_url": cells[32],
            "financial_need_proof_url": cells[33], "optional_supporting_url": cells[34], "certification": cells[35], "signature": cells[36], "signature_date": cells[37],
        }
        errors = [] if valid else ["Missing or invalid personal email address."]
        if duplicate:
            errors.append("Duplicate email within the supplied 2026 workbook.")
        parsed.append({"row_number": row_number, "source_record_id": f"sheet-row-{row_number}", "email": email or f"invalid-row-{row_number}@invalid.local", "submitted_at": cells[0] or None, "raw": dict(zip(unique_headers, cells)), "normalized": normalized, "status": "error" if not valid else "excluded" if duplicate else "committed", "exclusion_reason": errors[0] if errors else None, "errors": errors})

    valid_count = sum(item["status"] != "error" for item in parsed)
    duplicate_count = sum(item["status"] == "excluded" for item in parsed)
    print(json.dumps({"rows": len(parsed), "valid_email_rows": valid_count, "unique_applicants": valid_count - duplicate_count, "duplicate_rows_excluded": duplicate_count, "invalid_rows": len(parsed) - valid_count, "file_hash_prefix": digest[:12], "mode": "commit" if args.commit else "dry-run"}))
    if args.export:
        args.export.write_text(json.dumps({"file_hash": digest, "rows": parsed}, ensure_ascii=False), encoding="utf-8")
        print(json.dumps({"exported": True, "rows": len(parsed)}))
        return
    if not args.commit:
        return

    env = load_env()
    program = api("GET", "programs", env, params={"slug": "eq.name-your-need", "select": "id"}).json()[0]
    existing = api("GET", "import_jobs", env, params={"file_hash": f"eq.{digest}", "select": "id,status"}).json()
    if existing:
        job_id = existing[0]["id"]
    else:
        created = api("POST", "import_jobs", env, payload={"program_id": program["id"], "file_hash": digest, "mapping": {"version": 2, "source": "EFF supplied XLSX", "row_count": len(parsed)}, "status": "staged", "created_by": None}, prefer="return=representation").json()[0]
        job_id = created["id"]

    import_rows = [{"job_id": job_id, "row_number": item["row_number"], "raw_data": item["raw"], "normalized_data": item["normalized"], "action": "exclude" if item["status"] in ("excluded", "error") else "commit", "errors": item["errors"]} for item in parsed]
    legacy_rows = [{"source_system": "name_your_need_legacy_2026", "source_record_id": item["source_record_id"], "email": item["email"], "original_submitted_at": item["submitted_at"], "raw_data": item["raw"], "normalized_data": item["normalized"], "status": item["status"], "exclusion_reason": item["exclusion_reason"], "import_job_id": job_id} for item in parsed]
    for batch in chunks(import_rows):
        api("POST", "import_rows", env, payload=batch, params={"on_conflict": "job_id,row_number"}, prefer="resolution=merge-duplicates,return=minimal")
    for batch in chunks(legacy_rows):
        api("POST", "legacy_application_records", env, payload=batch, params={"on_conflict": "source_system,source_record_id"}, prefer="resolution=merge-duplicates,return=minimal")
    api("PATCH", "import_jobs", env, payload={"status": "committed", "committed_at": datetime.now().astimezone().isoformat()}, params={"id": f"eq.{job_id}"}, prefer="return=minimal")
    print(json.dumps({"committed": True, "job_id": job_id, "rows": len(parsed), "invitations_sent": 0}))


if __name__ == "__main__":
    main()
